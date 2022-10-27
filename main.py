from openpyxl import load_workbook
from obp_client import logger
from customers import create_customer_from_row
from accounts import create_account_with_customer_links_from_row
from transactions import create_transaction_from_row
from os import getcwd


data_source = f'{getcwd()}/resources/obp-data-import.xlsx'
wb = load_workbook(data_source)
customers = wb['Customers']
accounts = wb['Accounts']
transactions = wb['Transactions']


for row in customers.iter_rows(min_row=2, max_col=customers.max_column, max_row=customers.max_row):
	if row[0].value is None:
		continue
	logger.debug(f"start customer: {row[0].row}")
	create_customer_from_row(row)

for row in accounts.iter_rows(min_row=2, max_col=accounts.max_column, max_row=accounts.max_row):
	if row[0].value is None:
		continue
	logger.debug(f"start account: {row[0].row}")
	create_account_with_customer_links_from_row(row)

for row in transactions.iter_rows(min_row=2, max_col=transactions.max_column, max_row=transactions.max_row):
	if row[13].value is None:
		continue
	logger.debug(f"start transaction: {row[0].row}")
	create_transaction_from_row(row)

